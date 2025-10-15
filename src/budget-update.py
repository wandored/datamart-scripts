import glob

import pandas as pd
from openpyxl import load_workbook
from psycopg2 import sql
from psycopg2.errors import IntegrityError

from db_utils.dbconnect import DatabaseConnection
from db_utils.recreate_views import recreate_all_views


def read_budget_files():
    """
    read the budget csv files from ./downloads/budgets/ and conctaenate them into a single dataframe
    """
    year = input("Enter the year you want to update: ")
    csv_files = glob.glob("./downloads/budgets/*.csv")
    df = pd.concat((pd.read_csv(f) for f in csv_files), ignore_index=True)
    df["year"] = year
    df.rename(
        columns={
            "Location": "location",
            "glaccountnumber": "gl_number",
            "Account": "account_name",
            "Year": "year",
            "Period 01": "period_01",
            "Period 02": "period_02",
            "Period 03": "period_03",
            "Period 04": "period_04",
            "Period 05": "period_05",
            "Period 06": "period_06",
            "Period 07": "period_07",
            "Period 08": "period_08",
            "Period 09": "period_09",
            "Period 10": "period_10",
            "Period 11": "period_11",
            "Period 12": "period_12",
            "Period 13": "period_13",
        },
        inplace=True,
    )
    df.fillna(0, inplace=True)
    df_pivot = pd.melt(
        df,
        id_vars=["location", "gl_number", "account_name", "year"],
        var_name="period",
        value_name="amount",
    )
    df_pivot["period"] = df_pivot["period"].str.extract("(\d+)")
    df_pivot["year"] = df_pivot["year"].astype(int)
    df_pivot["period"] = df_pivot["period"].astype(int)

    # write to csv file
    df_pivot.to_csv(f"./output/{year}budgets.csv", index=False)

    return df_pivot


def process_budgets_xlsx(folder_path, output_path, year):
    xlsx_files = glob.glob(f"{folder_path}/*.xlsx")
    all_dfs = []

    for file in xlsx_files:
        # Load workbook and location (B2)
        wb = load_workbook(file, data_only=True)
        ws = wb.active
        location = ws["B2"].value

        # Load the actual sheet into pandas, skipping header rows if needed
        df = pd.read_excel(
            file, engine="openpyxl", skiprows=2
        )  # adjust skiprows as needed
        df["name"] = location

        # Split 'glaccountnumber' into 'gl_number' and 'account_name'
        split_cols = df["Account"].str.split(" - ", expand=True)
        df["gl_number"] = split_cols[0]
        df["account_name"] = split_cols[1]

        # Append to list
        all_dfs.append(df)

    # Combine all dataframes
    df = pd.concat(all_dfs, ignore_index=True)

    # Add year and rename columns
    df["year"] = year
    df.rename(
        columns={
            "Year": "year",
            "Period 01": "period_01",
            "Period 02": "period_02",
            "Period 03": "period_03",
            "Period 04": "period_04",
            "Period 05": "period_05",
            "Period 06": "period_06",
            "Period 07": "period_07",
            "Period 08": "period_08",
            "Period 09": "period_09",
            "Period 10": "period_10",
            "Period 11": "period_11",
            "Period 12": "period_12",
            "Period 13": "period_13",
        },
        inplace=True,
    )

    df.fillna(0, inplace=True)

    # Pivot the period columns into long format
    df_pivot = pd.melt(
        df,
        id_vars=["name", "gl_number", "account_name", "year"],
        value_vars=[
            "period_01",
            "period_02",
            "period_03",
            "period_04",
            "period_05",
            "period_06",
            "period_07",
            "period_08",
            "period_09",
            "period_10",
            "period_11",
            "period_12",
            "period_13",
        ],
        var_name="period",
        value_name="amount",
    )

    # Extract just the numeric portion of the period string (e.g., "01" from "period_01")
    df_pivot["period"] = df_pivot["period"].str.extract(r"(\d+)")

    # Ensure clean data before conversion
    df_pivot.dropna(subset=["period", "amount"], inplace=True)
    df_pivot["period"] = df_pivot["period"].astype(int)
    df_pivot["year"] = df_pivot["year"].astype(int)

    Save to CSV
    df_pivot.to_csv(f"{output_path}/{year}budgets.csv", index=False)
    """
    After writing file, it may be necessary to edit the location names manually
    """

    return


def write_to_database(cur, conn, engine, output_path, year):
    df = pd.read_csv(f"{output_path}/{year}budgets.csv", encoding="utf-7")
    df_restaurants = pd.read_sql("SELECT locationid, name FROM restaurants", db.conn)

    # Join on name
    df_merged = df.merge(df_restaurants, how="left", on="name")

    df_merged.rename(columns={"locationid": "location"}, inplace=True)

    table_name = "budgets"
    temp_table_name = "temp_table"
    try:
        df_merged.to_sql(temp_table_name, engine, if_exists="replace", index=False)
        update_query = sql.SQL(
            """
            INSERT INTO {table} (location, gl_number, account_name, year, period, amount)
            SELECT t.location, t.gl_number, t.account_name, t.year, t.period, t.amount
            FROM {temp_table} AS t
            ON CONFLICT (location, gl_number, account_name, year, period)
            DO UPDATE SET amount = EXCLUDED.amount
            """
        ).format(
            table=sql.Identifier(table_name),
            temp_table=sql.Identifier(temp_table_name),
        )
        cur.execute(update_query)
    except IntegrityError as e:
        print(e)
    finally:
        try:
            cur.execute("DROP TABLE IF EXISTS temp_table")
            conn.commit()
        except Exception as e:
            print("Error dropping temp table", e)
            conn.rollback()

    return


if __name__ == "__main__":
    with DatabaseConnection() as db:
        folder_path = "./downloads/budgets/"
        output_path = "./output/"
        year = input("Enter the year you want to update: ")

        process_budgets_xlsx(folder_path, output_path, year)
        # df = read_budget_files()
        write_to_database(db.cur, db.conn, db.engine, output_path, year)
        recreate_all_views(db.conn)
