import re
from datetime import date
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.utils import quote_sheetname

import pandas as pd

from config import Config
from dbconnect import DatabaseConnection
from toast_utils import extract_menu_items, get_access_token, get_response_data


def get_recipe_ingredient(cur, conn, engine) -> pd.DataFrame:
    cur.execute(
        """
        SELECT ri.concept,
            ri.menu_item,
            ri.recipe,
            ri.ingredient
        FROM recipe_ingredients ri
        WHERE ri.concept = 'Steakhouse' and ri.ingredient LIKE 'WINE %'
        ORDER BY ri.ingredient
        """
    )
    query = cur.fetchall()
    recipe_ingredient = pd.DataFrame(
        query,
        columns=["concept", "toast_item_name", "recipe_name", "purchase_item"],
    )

    cur.execute(
        """
        SELECT i.name FROM item i WHERE i.category2 = 'Wine' ORDER BY i.name
        """
    )
    query = cur.fetchall()
    items = pd.DataFrame(query, columns=["purchase_item"])

    # outer merge recipe_ingredient and items
    df = pd.merge(items, recipe_ingredient, how="outer", on="purchase_item")

    return df


def get_restaurant_names(guid_list, cur) -> pd.DataFrame:
    cur.execute(
        """
        SELECT r.name,
            r.toast_guid
        FROM restaurants r
        WHERE r.toast_guid = ANY(%s::uuid[])
        """,
        (guid_list,),
    )
    query = cur.fetchall()
    form_template = pd.DataFrame(
        query,
        columns=["location_name", "location_guid"],
    )

    return form_template


def get_menu_list(api_access_url, token, guid_list):
    df_menus = pd.DataFrame()
    for guid in guid_list:
        extracted_data = []
        url = api_access_url + "/menus/v2/menus"
        headers = {
            "Toast-Restaurant-External-ID": guid,
            "Authorization": f"Bearer {token}",
        }
        menus = get_response_data(url, headers)

        for menu in menus:
            menu_id = menu.get("guid", "")
            if menu_id != "f3e6905c-8e29-4f65-a7dc-55336b1a544d":
                continue
            menu_name = menu.get("name", "")
            for group in menu.get("menuGroups", []):
                extracted_data.extend(
                    extract_menu_items(guid, menu_id, menu_name, group)
                )
        if extracted_data:
            df_extracted = pd.DataFrame(extracted_data)
            df_menus = pd.concat([df_menus, df_extracted], ignore_index=True)

    return df_menus


def extract_correct_pos_name(row):
    code_map = {
        "NYP-Atlanta": "A",
        "NYP-Boca Raton": "B",
        "NYP-Myrtle Beach": "M",
        "Chophouse 47": "G",
        "Chophouse-NOLA": "L",
        "BT Prime Rib": "T",
    }

    try:
        location_code = row["location_name"] if "location_name" in row else None
    except (ValueError, TypeError):
        return None

    pos_name = row["pos_name"] if "pos_name" in row else ""

    if (
        not location_code
        or not pos_name
        or not isinstance(pos_name, str)
        or pos_name.strip().lower() == "nan"
    ):
        return None

    prefix = code_map.get(location_code)
    if not prefix:
        return None

    # Adjust regex to match codes with the correct prefix
    regex = rf"\b{prefix}\d{{3,4}}\b"
    match = re.search(regex, pos_name)
    if match:
        return match.group(0)

    return None


def create_excel_for_location(combined_df, location, today_date):
    filename = f"/home/wandored/Sync/ReportData/Wine_Forms/{location}_wine_template_{today_date}.xlsx"

    # Save combined_df to Sheet1 ("Data")
    combined_df.to_excel(filename, sheet_name="Data", index=False, engine="openpyxl")

    # Load workbook and get Data sheet
    wb = load_workbook(filename)
    data_sheet = wb["Data"]

    # Add Excel table to Data sheet
    max_row = len(combined_df) + 1  # +1 for header
    data_table = Table(displayName="DataTable", ref=f"A1:E{max_row}")
    data_table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    data_sheet.add_table(data_table)

    # Create Wine Form sheet
    table_sheet = wb.create_sheet(title="Wine Form")

    # Define your Wine Form headers
    wine_form_headers = [
        "Action",
        "R365 Purchase Item",
        "Toast Button",
        "Bin Number",
        "Menu Description",
        "Book Page",
        "Vintage Required",
        "Bottle Size",
        "Price Paid",
        "Target Price 33%",
        "Menu Price",
        "Cost %",
        "New Vintage",
        "Vendor Item Number",
        "Comments",
    ]

    # Write headers into Wine Form
    for col_num, header in enumerate(wine_form_headers, start=1):
        table_sheet.cell(row=1, column=col_num).value = header

    # Define named range for purchase_items
    max_row = len(combined_df) + 1
    purchase_item_range = f"{quote_sheetname('Data')}!$A$2:$A${max_row}"
    dn = DefinedName("purchase_items", attr_text=purchase_item_range)
    wb.defined_names.add(dn)

    # Add Data Validation dropdown for "R365 Purchase Item" (column B)
    dv_purchase = DataValidation(
        type="list",
        formula1="=purchase_items",  # A2:A{N}
        allow_blank=True,
        showDropDown=False,
    )
    table_sheet.add_data_validation(dv_purchase)
    dv_purchase.add(f"B2:B30")  # Apply dropdown to B2 through B30

    # Data validation for Action column (Column A)
    dv_action = DataValidation(
        type="list",
        formula1='"Add,Delete,Vintage,Other"',
        allow_blank=True,
        showDropDown=False,
    )
    table_sheet.add_data_validation(dv_action)
    dv_action.add("A2:A30")

    # Data validation for Vintage column (Column G)
    dv_action = DataValidation(
        type="list",
        formula1='"Yes,No"',
        allow_blank=True,
        showDropDown=False,
    )
    table_sheet.add_data_validation(dv_action)
    dv_action.add("G2:G30")

    # Add Excel table to Wine Form sheet
    form_table = Table(displayName="WineFormTable", ref="A1:O30")  # 15 columns
    form_table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    table_sheet.add_table(form_table)

    # Set Excel formulas in rows
    for row in range(2, 31):
        # Toast Button from R365 Purchase Item (Column C)
        table_sheet[f"C{row}"].value = (
            f'=IFERROR(VLOOKUP(B{row}, Data!A:B, 2, FALSE), "")'
        )

        # Bin Number from R365 Purchase Item (Column D)
        table_sheet[f"D{row}"].value = (
            f'=IFERROR(VLOOKUP(B{row}, Data!A:E, 5, FALSE), "")'
        )

        # Target Price 33% (Column J) = Price Paid / 0.33
        table_sheet[f"J{row}"].value = f'=IFERROR(I{row}/0.33, "")'

        # Cost % (Column L) = Price Paid / Menu Price
        table_sheet[f"L{row}"].value = f'=IF(OR(K{row}=0, K{row}=""), 0, I{row}/K{row})'

        # Apply Accounting format
        table_sheet[f"I{row}"].number_format = (
            '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'  # Price Paid
        )
        table_sheet[f"J{row}"].number_format = (
            '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'  # Target Price
        )
        table_sheet[f"K{row}"].number_format = (
            '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'  # Menu Price
        )

        # Apply Percentage format with 1 decimal place
        table_sheet[f"L{row}"].number_format = "0.0%"

    wb.save(filename)
    print(f"Saved Excel file for location {location}.")


def create_worksheets(df):
    today_date = date.today().strftime("%Y-%m-%d")
    locations = [
        "NYP-Atlanta",
        "NYP-Boca Raton",
        "NYP-Myrtle Beach",
        "Chophouse 47",
        "Chophouse-NOLA",
        "BT Prime Rib",
    ]
    columns_to_drop = [
        "location_guid",
        "menu_id",
        "menu_name",
        "menu_group_name",
        "pos_name",
        "kitchen_name",
        "concept",
    ]
    df_cleaned = df.drop(columns=columns_to_drop)
    df_cleaned.to_excel(
        "/home/wandored/Sync/ReportData/Wine_Forms/Steakhouse_wine_links.xlsx",
        index=False,
    )
    toast_orphans = df_cleaned[
        df_cleaned["location_name"].isna() | (df_cleaned["location_name"] == "")
    ]

    # import all wine purchase items
    purchase_items = pd.read_csv(
        "./downloads/PurchaseItems.csv",
        usecols=["Name"],
        encoding="ISO-8859-1",
    ).rename(columns={"Name": "purchase_item"})

    # Strip spaces
    purchase_items["purchase_item"] = purchase_items["purchase_item"].str.strip()

    # Filter wines
    wine_purchase_items = purchase_items[
        purchase_items["purchase_item"].str.startswith("WINE ")
    ]
    r365_orphans = wine_purchase_items[
        ~wine_purchase_items["purchase_item"].isin(df_cleaned["purchase_item"])
    ]

    for location in locations:
        location_rows = df_cleaned[df_cleaned["location_name"] == location]
        # clean all wines not used in location
        other_location_rows = df_cleaned[df_cleaned["location_name"] != location].copy()
        other_location_rows["location_name"] = ""
        other_location_rows["bin_number"] = ""
        unique_other_rows = other_location_rows.drop_duplicates(
            subset=["toast_item_name"]
        )
        unique_other_rows = other_location_rows[
            ~other_location_rows["toast_item_name"].isin(
                location_rows["toast_item_name"]
            )
        ]
        combined_df = pd.concat(
            [location_rows, unique_other_rows, toast_orphans, r365_orphans],
            ignore_index=True,
        )
        combined_df = combined_df.drop_duplicates(subset="toast_item_name")
        if not combined_df.empty:
            create_excel_for_location(combined_df, location, today_date)


def main():
    # Authenticate and get access token
    guid_list = Config.LOCATION_GUID_LIST
    api_access_url = "https://ws-api.toasttab.com"
    token = get_access_token(api_access_url)
    if not isinstance(token, dict):
        raise TypeError("Expected token to be a dictionary")
    access_token = token.get("accessToken", "")

    wine_menus = get_menu_list(api_access_url, access_token, guid_list)

    with DatabaseConnection() as db:
        steakhouse_names = get_restaurant_names(guid_list, db.cur)
        steakhouse_wines = pd.merge(
            wine_menus,
            steakhouse_names,
            how="left",
            on="location_guid",
        )
        # Split out bin numbers
        steakhouse_bins = steakhouse_wines.copy()
        steakhouse_bins["bin_number"] = steakhouse_wines.apply(
            extract_correct_pos_name, axis=1
        )
        recipe_ingredients = get_recipe_ingredient(db.cur, db.conn, db.engine)
        df = pd.merge(
            recipe_ingredients, steakhouse_bins, how="outer", on="toast_item_name"
        )
        create_worksheets(df)

        df_pivot = df.pivot_table(
            index=["purchase_item", "recipe_name", "toast_item_name"],
            columns="location_name",
            values="bin_number",
            aggfunc="first",
        )
        df_pivot.to_csv("./output/steakhouse_wine_links.csv", index=True)
        print(df_pivot)


if __name__ == "__main__":
    main()
